"""
Excel report generator for evaluation results.

Generates Excel files with multiple sheets for comprehensive analysis.
"""
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_excel_report(results: Dict[str, Any], output_path: Path, conversation: List[Dict[str, str]] = None) -> None:
    """
    Generate an Excel report from evaluation results.
    
    Args:
        results: Evaluation results dictionary
        output_path: Path where Excel file should be saved
        conversation: Original conversation data
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. Skipping Excel generation.")
        print("Install with: pip install openpyxl")
        return
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create sheets
    create_summary_sheet(wb, results)
    create_details_sheet(wb, results, conversation)
    create_raw_data_sheet(wb, results)
    
    # Save Excel file
    excel_path = output_path.with_suffix('.xlsx')
    wb.save(excel_path)


def create_summary_sheet(wb: Workbook, results: Dict[str, Any]) -> None:
    """Create summary overview sheet."""
    ws = wb.create_sheet("Summary")
    
    # Title
    ws['A1'] = "Evaluation Summary"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    ws.merge_cells('A1:B1')
    
    # File info
    row = 3
    ws[f'A{row}'] = "File:"
    ws[f'B{row}'] = results.get('file', 'Unknown')
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Timestamp:"
    ws[f'B{row}'] = results.get('timestamp', 'Unknown')
    ws[f'A{row}'].font = Font(bold=True)
    
    # Metrics count
    row += 2
    evaluator_count = len(results.get('evaluator_results', {}).get('results', {}))
    literature_count = len(results.get('literature_results', {}).get('results', {}))
    
    ws[f'A{row}'] = "Evaluator Metrics:"
    ws[f'B{row}'] = evaluator_count
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Literature Metrics:"
    ws[f'B{row}'] = literature_count
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Total Metrics:"
    ws[f'B{row}'] = evaluator_count + literature_count
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True, color="667EEA")
    
    # Status
    row += 2
    errors = results.get('errors', [])
    if errors:
        ws[f'A{row}'] = "Status:"
        ws[f'B{row}'] = f"⚠ {len(errors)} Error(s)"
        ws[f'B{row}'].font = Font(color="DC2626")
        
        row += 1
        ws[f'A{row}'] = "Errors:"
        ws[f'A{row}'].font = Font(bold=True)
        for error in errors:
            row += 1
            ws[f'B{row}'] = f"• {error}"
    else:
        ws[f'A{row}'] = "Status:"
        ws[f'B{row}'] = "✓ Success"
        ws[f'B{row}'].font = Font(color="16A34A")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50


def create_details_sheet(wb: Workbook, results: Dict[str, Any], conversation: List[Dict[str, str]] = None) -> None:
    """Create detailed turn-by-turn analysis sheet."""
    ws = wb.create_sheet("Turn-by-Turn Analysis")
    
    # Collect metrics and utterances
    metric_names = []
    utterances_data = []
    
    # From evaluator results
    if results.get('evaluator_results') and results['evaluator_results'].get('results'):
        for metric_name, metric_info in results['evaluator_results']['results'].items():
            if metric_name not in metric_names:
                metric_names.append(metric_name)
            
            if metric_info.get('per_utterance'):
                for i, utt in enumerate(metric_info['per_utterance']):
                    if i >= len(utterances_data):
                        utterances_data.append({})
                    if utt.get('metrics'):
                        utterances_data[i].update(utt['metrics'])
    
    # From literature results
    if results.get('literature_results') and results['literature_results'].get('results'):
        for metric_name, metric_info in results['literature_results']['results'].items():
            if metric_name not in metric_names:
                metric_names.append(metric_name)
            
            if metric_info.get('per_utterance'):
                for i, utt in enumerate(metric_info['per_utterance']):
                    if i >= len(utterances_data):
                        utterances_data.append({})
                    if utt.get('metrics'):
                        utterances_data[i].update(utt['metrics'])
    
    # Create headers
    headers = ['#', 'Speaker', 'Utterance']
    for metric in metric_names:
        headers.append(metric)
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, utt_metrics in enumerate(utterances_data, 2):
        # Turn number
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        
        # Speaker and text
        if conversation and (row_idx - 2) < len(conversation):
            conv_item = conversation[row_idx - 2]
            ws.cell(row=row_idx, column=2, value=conv_item.get('speaker', ''))
            ws.cell(row=row_idx, column=3, value=conv_item.get('text', ''))
        
        # Metrics
        for col_idx, metric_name in enumerate(metric_names, 4):
            score = utt_metrics.get(metric_name, {})
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if score:
                if score.get('type') == 'numerical':
                    cell.value = score.get('value', 0)
                    cell.number_format = '0.00'
                    # Color code based on value
                    ratio = score.get('value', 0) / score.get('max_value', 1)
                    if ratio >= 0.8:
                        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                    elif ratio >= 0.5:
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                elif score.get('type') == 'categorical':
                    cell.value = score.get('label', 'N/A')
            else:
                cell.value = '-'
                cell.font = Font(color="94A3B8")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 60
    for col_idx in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Freeze panes (freeze first row and first 3 columns)
    ws.freeze_panes = 'D2'


def create_raw_data_sheet(wb: Workbook, results: Dict[str, Any]) -> None:
    """Create sheet with raw JSON data for reference."""
    ws = wb.create_sheet("Raw Data")
    
    import json
    
    ws['A1'] = "Raw JSON Data"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Pretty print JSON
    json_str = json.dumps(results, indent=2)
    
    # Split into lines and write
    for row_idx, line in enumerate(json_str.split('\n'), 3):
        ws.cell(row=row_idx, column=1, value=line)
    
    ws.column_dimensions['A'].width = 100
